from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Border,
    Side,
    Alignment
)

class ExcelExporter:

    def __init__(self, carpeta_salida):

        self.carpeta = Path(carpeta_salida)

        self.carpeta.mkdir(
            parents=True,
            exist_ok=True
        )

        self.carpeta_dependencias = (
            self.carpeta / "Dependencias"
        )

        self.carpeta_dependencias.mkdir(
            exist_ok=True
        )

    # ---------------------------------------------------------

    def exportar(self, dataframe):

        self._exportar_general(dataframe)

        self._exportar_dependencias(dataframe)


        # ---------------------------------------------------------

    def _exportar_general(self, dataframe):

        archivo = self.carpeta / "Consolidado_General.xlsx"

        df = dataframe.copy()

        df = df.rename(
            columns={
                "actividad": "Actividad",
                "documento": "Documento",
                "nombre": "Nombre",
                "dependencia": "Dependencia",
                "cargo": "Cargo",
                "tipo_vinculacion": "Tipo vinculación",
                "sesiones_totales": "Sesiones totales",
                "sesiones_realizadas": "Sesiones realizadas",
                "sesiones_asistidas": "Sesiones asistidas",
                "sesiones_falladas": "Sesiones falladas",
                "detalle_inasistencias": "Sesiones no asistidas",
                "porcentaje_asistencia": "% asistencia",
            }
        )

        df.to_excel(
            archivo,
            index=False
        )

        self._formatear_excel(archivo)

    # ---------------------------------------------------------

    def _exportar_dependencias(self, dataframe):

        for dependencia in sorted(
            dataframe["dependencia"].dropna().unique()
        ):

            df = dataframe[
                dataframe["dependencia"] == dependencia
            ].copy()

            df = df.rename(
                columns={
                    "actividad": "Actividad",
                    "documento": "Documento",
                    "nombre": "Nombre",
                    "dependencia": "Dependencia",
                    "cargo": "Cargo",
                    "tipo_vinculacion": "Tipo vinculación",
                    "sesiones_totales": "Sesiones totales",
                    "sesiones_realizadas": "Sesiones realizadas",
                    "sesiones_asistidas": "Sesiones asistidas",
                    "sesiones_falladas": "Sesiones falladas",
                    "detalle_inasistencias": "Sesiones no asistidas",
                    "porcentaje_asistencia": "% asistencia",
                }
            )

            nombre = "".join(
                c
                for c in dependencia
                if c not in '\\/:*?"<>|'
            ).strip()

            archivo = (
                self.carpeta_dependencias
                / f"{nombre}.xlsx"
            )

            df.to_excel(
                archivo,
                index=False
            )

            self._formatear_excel(archivo)
    # ---------------------------------------------------------

    def _formatear_excel(self, archivo):

        libro = load_workbook(archivo)

        hoja = libro.active
        hoja.freeze_panes = "A2"
        hoja.sheet_view.showGridLines = False

        encabezado = PatternFill(
            fill_type="solid",
            start_color="1F4E78",
            end_color="1F4E78"
        )

        verde = PatternFill(
            fill_type="solid",
            start_color="C6EFCE",
            end_color="C6EFCE"
        )

        rojo = PatternFill(
            fill_type="solid",
            start_color="FFC7CE",
            end_color="FFC7CE"
        )
        borde = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        centrado = Alignment(
            horizontal="center",
            vertical="center"
        )

        izquierda = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

        #
        # Encabezados
        #
        for celda in hoja[1]:

            celda.fill = encabezado

            from openpyxl.styles import Font

            celda.font = Font(
                bold=True,
                color="FFFFFF"
            )

            celda.border = borde
            celda.alignment = centrado

        #
        # Ajustar ancho automáticamente
        #
        for columna in hoja.columns:

            longitud = 0

            letra = columna[0].column_letter

            for celda in columna:

                try:

                    longitud = max(
                        longitud,
                        len(str(celda.value))
                    )

                except Exception:

                    pass

            hoja.column_dimensions[letra].width = min(
                longitud + 2,
                60
            )
        for fila in hoja.iter_rows(min_row=2):

            for celda in fila:

                celda.border = borde

                encabezado_columna = hoja.cell(
                    row=1,
                    column=celda.column
                ).value

                if encabezado_columna in (
                    "Documento",
                    "Sesiones totales",
                    "Sesiones realizadas",
                    "Sesiones asistidas",
                    "Sesiones falladas",
                    "% asistencia",
                ):

                    celda.alignment = centrado

                else:

                    celda.alignment = izquierda
            

        #
        # Pintar porcentaje de asistencia
        #
        porcentaje_col = None

        for celda in hoja[1]:

            if celda.value == "% asistencia":

                porcentaje_col = celda.column

                break

        if porcentaje_col is not None:

            for fila in range(2, hoja.max_row + 1):

                celda = hoja.cell(
                    row=fila,
                    column=porcentaje_col
                )

                valor = celda.value
                celda.number_format = "0%"

                if valor is None:
                    continue

                try:

                    if float(valor) >= 0.80:

                        celda.fill = verde

                    else:

                        celda.fill = rojo

                except Exception:

                    pass
        hoja.auto_filter.ref = hoja.dimensions
        libro.save(archivo)