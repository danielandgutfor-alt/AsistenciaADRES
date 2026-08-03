"""
Módulo encargado de leer el cronograma de asistencia.
"""

from pathlib import Path

from openpyxl import load_workbook


class LectorCronograma:

    def __init__(self, ruta_archivo: Path):

        self.ruta_archivo = ruta_archivo
        self.libro = None

    def cargar(self):

        if not self.ruta_archivo.exists():

            raise FileNotFoundError(
                f"No existe el archivo:\n{self.ruta_archivo}"
            )

        self.libro = load_workbook(
            filename=self.ruta_archivo,
            data_only=False
        )

    def obtener_hojas(self):

        if self.libro is None:

            raise Exception(
                "Primero debes cargar el cronograma."
            )

        return self.libro.sheetnames
